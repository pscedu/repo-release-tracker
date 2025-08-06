#code from yesterday that gets the latest releases of the singularity repos :

# Ok First lets add in our code first and clean it up. Then lets add the stuff for sending emails

#import libraries: PyGithub for accessing github API and pandas for dataframes
from pydoc import html
import sys
print(sys.executable)

from github import Github
import pandas as pd

#libaries for creating spreadsheets
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from datetime import datetime



#create github instance and gets repo:
g = Github("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")    # says can do 60 requests per hour but there are more than 60.


#get the names of the repos we want to check:

stem_repos = ['stride', 'nanoplot','star-fusion','filtlong','porechop','anvio','funannotate','fastq-tools','meme-suite','braker2','rust','guppy','guppy-gpu','bsmap','salmon','rnaview','bioformats2raw','raw2ometiff','flash','blat','bedops','genemark-es','augustus','checkm','ncview','bowtie2','asciigenome','fastqc','sra-toolkit','gatk','hmmer','bcftools','raxml','spades','busco','samtools','bedtools','bamtools','fastani','phylip-suite','blast','viennarna','cutadapt','bismark','star','prodigal','bwa','picard','hisat2','abyss','octave','tiger','gent','methylpy','fasttree','vcf2maf','htslib','kraken2','aspera-connect','trimmomatic']
utilities_repos = ['hashdeep','dua','vim','timewarrior','libtiff-tools','wordgrinder','shellcheck','pandiff','rich-cli','jq','jp','lowcharts','btop','aws-cli','cwltool','circos','glances','fdupes','graphviz','browsh','hyperfine','dust','gnuplot','pandoc','mc','bat','flac','visidata','octave','ncdu','lazygit','asciinema','ffmpeg','imagemagick','rclone']

#now lets go through them and extract the names, version, and years
#starting with stem and then utilities repo:
print("starting")
stemRepo = []
utilityRepo = []
for repoName in stem_repos:
    try:
        repo = g.get_repo("pscedu/singularity-"+repoName)
        latestRelease = repo.get_latest_release()
        release = latestRelease.title or latestRelease.tag_name
        date = str(latestRelease.published_at)[0:10]
    
        stemRepo.append([repoName, release, date])
        print("finished " + repoName +"\n")
    except Exception as e:
        stemRepo.append([repoName, "No release", "N/A"])
#now for utility repos:
for repoName in utilities_repos:
    try:
        repo = g.get_repo("pscedu/singularity-"+repoName)
        latestRelease = repo.get_latest_release()
        release = latestRelease.title or latestRelease.tag_name
        date = str(latestRelease.published_at)[0:10]
    
        utilityRepo.append([repoName, release, date])
    except Exception as e:
        utilityRepo.append([repoName, "No release", "N/A"])

#Now we need to figure out how to sort them by category. 





#now lets make the dataframes for the repos:
df_stem = pd.DataFrame(stemRepo, columns = ["Repo Name", "Latest Version", "Date"])
df_utility = pd.DataFrame(utilityRepo, columns = ["Repo Name", "Latest Version", "Date"])
print("finished making dataframes")

#now lets combine the dataframes into one dataframe and create new category column:

df_stem["Category"] = "STEM" #adds category column
df_utility["Category"] = "Utility"

combined_df = pd.concat([df_stem, df_utility], ignore_index=True) #combines without losing duplicates

category_map = combined_df.groupby("Repo Name")["Category"].apply(lambda x: "Both" if len(set(x)) > 1 else x.iloc[0])
combined_df["Category"] = combined_df["Repo Name"].map(category_map) #removes duplicates and adds updated category 

combined_df = combined_df.drop_duplicates(subset=["Repo Name"]) #removes duplicates based on repo name


# ok now let's sort these by date(descending order). and then rearrange the columns

combined_df["Date"] = pd.to_datetime(combined_df["Date"], errors="coerce")
sorted_combined_df = combined_df.sort_values(by="Date", ascending=False)



# now lets rearrange the columns to have  category first, then repo name, then latest version, and then date
sorted_combined_df = sorted_combined_df[["Category", "Repo Name", "Latest Version", "Date"]]
print(sorted_combined_df)

# now lets create the spread sheet

wb = Workbook()
ws = wb.active
ws.title = "PSC Repo Versions"

# create header row
header = ["Category", "Repo Name", "Latest Version", "Date"]
ws.append(header)

# add data rows while adding links
for _, row in sorted_combined_df.iterrows():
    excel_row = ws.max_row + 1 #the row we are currently on

    #column 1: category
    ws.cell(row=excel_row, column=1, value=row["Category"])

    #column 2: repo name with link
    repo_cell = ws.cell(row=excel_row, column=2, value=row['Repo Name'])
    repo_link = f"https://github.com/pscedu/singularity-{row['Repo Name']}"
    repo_cell.hyperlink = repo_link
    repo_cell.font = Font(underline="single", color="0000FF")  # blue and underlined
    
    #column 3: latest version with link
    version_cell = ws.cell(row=excel_row, column=3, value=row["Latest Version"])
    version_link = f"https://github.com/pscedu/singularity-{row['Repo Name']}/releases"
    version_cell.hyperlink = version_link
    version_cell.font = Font(underline="single", color="0000FF")  # blue and underlined

    # Column 4: date
    ws.cell(row=excel_row, column=4, value=row["Date"])
    



# now we'll name it and save it:

today = datetime.now().strftime("%Y-%m-%d")
file_name = f"github_releases_{today}.xlsx"

# save the workbook
wb.save("Repo_Versions.xlsx")

print(f"Spreadsheet saved as {file_name}")
